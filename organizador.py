import pandas as pd
import openpyxl
from tkinter import *
import tkinter.messagebox as messagebox
from tkinter import ttk
from tkinter import filedialog


class MainExcel():
    def __init__(self):
        pass

    def run (self):
    # Leitura do CSV
        #arquivo = input("Digite o caminho do arquivo gerado pelo ServiceNow >>>")
        try:
            file_path = filedialog.askopenfilename(title="Digite o caminho do arquivo gerado pelo ServiceNow >>>", filetypes=[("csv", "*.csv")])
            dados = pd.read_csv(file_path, encoding="latin-1")
    # Criação da pasta de trabalho e planilha
            job = openpyxl.Workbook()
            planilha = job.active

    # Escrita dos cabeçalhos
            for i, colunas in enumerate(dados.columns):
                planilha.cell(row=1, column=i+1).value = colunas

    # Preenchimento dos dados
            for row_num, row in dados.iterrows():
                for col_num, colunas in enumerate(row):
                    planilha.cell(row=row_num+2, column=col_num+1).value = colunas

    # Salvamento do arquivo Excel
            job.save('arquivo_organizado.xlsx')
            print("Arquivo Excel organizado com sucesso!")
            messagebox.showinfo('Sucesso, dados organizados com sucesso!')
        except:
            messagebox.showerror("Erro", f"Ocorreu um erro ao carregar o arquivo csv:")
             
if __name__ == "__main__":
    run = MainExcel()
    run.run()
